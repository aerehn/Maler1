from openpyxl import load_workbook
from tkinter import *
import pandas as pd
import numpy as np
from openpyxl import Workbook
def moro(c=0,debug=True):
    if(debug==True):
        print(c)

def ajaArvot(lahdeRivit=-1, kohdeRivi=32, lahde="export_Tuotteiden_vienti_XLSX_2021-10-09_14-42-15.xlsx", kohde="SOK Käyttötavaroiden erätuotelomake (muut käyttötavarat) v2 33 (version 1) (version 1)_ROSTERi.xlsx",console=0,debug=True):
    def write(*message, end="\n", sep=" "):
        text = ""
        for item in message:
            text += "{}".format(item)
            text += sep
        text += end
        console.insert(INSERT, text)


    #function for inserting data in column form
    def writeColumn(columnList,sheet,targetRow,targetCol,sourceRows,maxLen=0):
        iterator = 0
        for i in range(sourceRows[0]-2,sourceRows[1]-1):
            sheet[targetCol+str(iterator+targetRow)] = columnList[i]
            iterator = iterator + 1

    def writeVanhatTuotteet(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name("Vanhat tuotteet - Old articles")
        #sku
        writeColumn(source['sku'], sheet, targetRow + offset, "A", sourceRows =sourceRows)
        #
        writeColumn(source['pitka_tuotenimi-fi_FI'], sheet, targetRow + offset, "C", sourceRows =sourceRows)

    #offset=32
    def writeUutuudet(workbook, targetRow, source, sourceRows):
        offset = 0

        sheet = workbook.get_sheet_by_name('1. Uutuudet - New articles')
        sourceCols = ['sku','pitka_tuotenimi-fi_FI','pitka_tuotenimi-fi_FI','tuotemerkki','kappalettalavalla','tuotteen_nettopaino','pituus','leveys']
        targetCols = ["N","O","P","T","CJ","DQ","DU","DW"]

        writeColumn(source['sku'], sheet, targetRow + offset, "AB", sourceRows =sourceRows)
        writeColumn(source['etiketin_lisateksti_25-fi_FI'], sheet, targetRow + offset, "EN", sourceRows=sourceRows)
        writeColumn(source['jmpaketissa-unit'], sheet, targetRow + offset, "AJ", sourceRows=sourceRows)
        writeColumn(source['koko'], sheet, targetRow + offset, "BR", sourceRows =sourceRows)
        writeColumn(source['korkeus'], sheet, targetRow + offset, "DY", sourceRows =sourceRows)
        writeColumn(source['kplpaketissa'], sheet, targetRow + offset, "AO", sourceRows =sourceRows)
        writeColumn(source['leveys'], sheet, targetRow + offset, "DW", sourceRows =sourceRows)
        writeColumn(source['myyntierana_hyllytettava'], sheet, targetRow + offset, "EC", sourceRows=sourceRows)
        writeColumn(source['pituus'], sheet, targetRow + offset, "DU", sourceRows =sourceRows)
        writeColumn(source['raaka_aine_materiaali'], sheet, targetRow + offset, "BW", sourceRows =sourceRows)
        writeColumn(source['savy_vari-fi_FI'], sheet, targetRow + offset, "BT", sourceRows =sourceRows)
        writeColumn(source['tekninenvarinumero'], sheet, targetRow + offset, "BS", sourceRows=sourceRows)
        writeColumn(source['tilausnumero'], sheet, targetRow + offset, "AQ", sourceRows =sourceRows)
        writeColumn(source['tullikoodi_nimike'], sheet, targetRow + offset, "ET", sourceRows=sourceRows)
        writeColumn(source['tuotekuvaus_markkinointiteksti-fi_FI'], sheet, targetRow + offset, "EV", sourceRows =sourceRows)
        writeColumn(source['tuotemerkki'], sheet, targetRow + offset, "T", sourceRows =sourceRows)
        writeColumn(source['hyllynreuna_25-fi_FI'], sheet, targetRow + offset, "EL", sourceRows=sourceRows)
        writeColumn(source['hyllynreuna_25-sv_SE'], sheet, targetRow + offset, "EP", sourceRows=sourceRows)
        writeColumn(source['tuotenimi_40_merkkia-en_GB'], sheet, targetRow + offset, "R", sourceRows=sourceRows)
        writeColumn(source['tuotenimi_40_merkkia-fi_FI'], sheet, targetRow + offset, "P", maxLen=40,
                    sourceRows=sourceRows)
        writeColumn(source['tuotteen_nettopaino'], sheet, targetRow + offset, "DQ", sourceRows =sourceRows)
        writeColumn(source['tuotteen_bruttopaino'], sheet, targetRow + offset, "DO", sourceRows =sourceRows)
        writeColumn(source['tuotteen_perusmaarayksiko'], sheet, targetRow + offset, "X", sourceRows=sourceRows)

        #sheet['AS39']="Kyllä / Yes"


    def writeToimitusyks(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name('2. Toimitusyks. -  Deliv. units')
        #writables get written
        writeColumn(source['korkeus'], sheet, targetRow + offset, "M", sourceRows=sourceRows)
        writeColumn(source['lavakorkeus'], sheet, targetRow + offset, "BV", sourceRows=sourceRows)
        writeColumn(source['lavanbruttopaino'], sheet, targetRow + offset, "BX", sourceRows=sourceRows)
        writeColumn(source['lavanleveys'], sheet, targetRow + offset, "BT", sourceRows=sourceRows)
        writeColumn(source['lavanpituus'], sheet, targetRow + offset, "BR", sourceRows=sourceRows)
        writeColumn(source['leveys'], sheet, targetRow + offset, "K", sourceRows=sourceRows)
        writeColumn(source['paketinbruttopaino'], sheet, targetRow + offset, "AR", sourceRows=sourceRows)
        writeColumn(source['paketti_korkeus'], sheet, targetRow + offset, "AP", sourceRows=sourceRows)
        writeColumn(source['paketti_leveys'], sheet, targetRow + offset, "AN", sourceRows=sourceRows)
        writeColumn(source['paketti_syvyys'], sheet, targetRow + offset, "AL", sourceRows=sourceRows)
        writeColumn(source['pituus'], sheet, targetRow + offset, "I", sourceRows =sourceRows)
        writeColumn(source['tuotteen_bruttopaino'], sheet, targetRow + offset, "O", sourceRows =sourceRows)

    def writeNimet(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name('3. Nimet - Names')
        writeColumn(source['pitka_tuotenimi-en_GB'], sheet, targetRow + offset, "U", sourceRows=sourceRows)
        writeColumn(source['pitka_tuotenimi-fi_FI'], sheet, targetRow + offset, "C", sourceRows=sourceRows)
        writeColumn(source['pitka_tuotenimi-sv_SE'], sheet, targetRow + offset, "M", sourceRows=sourceRows)
        writeColumn(source['tuotenimi_40_merkkia-sv_SE'], sheet, targetRow + offset, "O", sourceRows=sourceRows)





    """
    def writeToimitusyks(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name('2. Toimitusyks. -  Deliv. units')
    def writeToimitusyks(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name('2. Toimitusyks. -  Deliv. units')
    def writeToimitusyks(workbook, targetRow, source, sourceRows):
        offset = 0
        sheet = workbook.get_sheet_by_name('2. Toimitusyks. -  Deliv. units')
    """

    """
    start_time = time.time()
    sok_pim = load_workbook(filename="PIM - SOK -infoa.xlsx")
    SOK_Ke = load_workbook(filename="SOK Käyttötavaroiden erätuotelomake (muut käyttötavarat) v2 33 (version 1) (version 1)_ROSTERi.xlsx")
    print(sok_pim.sheetnames)
    print(SOK_Ke.sheetnames)
    print("Loading took %s seconds" % (time.time()-start_time))
    start_time = time.time()
    sheet2 = SOK_Ke.active
    sheet = sok_pim.active
    sheet['A15'] = "Well hello there!"
    sheet2['A40'] = 'moro'
    sheet2['A41'] = sheet['ER1'].value
    print("Modifications took %s seconds" % (time.time()-start_time))
    start_time = time.time()
    SOK_Ke.save(filename="SOK Käyttötavaroiden erätuotelomake (muut käyttötavarat) v2 33 (version 1) (version 1)_ROSTERi.xlsx")
    sok_pim.save(filename="PIM - SOK -infoa.xlsx")
    print("Saving took %s seconds" % (time.time()-start_time))
    """

    moro("Loading source Workbook...", debug)
    lahdeDF = pd.read_excel(lahde)
    moro(lahdeDF,debug)
    moro("Loading target Workbook...", debug)
    SOK_Ke = load_workbook(filename=kohde)
    moro(SOK_Ke.sheetnames,debug)

    sheet2 = SOK_Ke.active
    moro("Writing sheets...", debug)
    if lahdeRivit==-1:
        shape = lahdeDF.shape
        lahdeRivit=[2,shape[0]+1]
    #writeVanhatTuotteet(SOK_Ke, targetRow=kohdeRivi, source=lahdeDF,sourceRows = lahdeRivit)
    writeUutuudet(SOK_Ke, targetRow=kohdeRivi, source=lahdeDF,sourceRows = lahdeRivit)
    writeToimitusyks(SOK_Ke, targetRow=kohdeRivi, source=lahdeDF,sourceRows = lahdeRivit)
    writeNimet(SOK_Ke, targetRow=kohdeRivi, source=lahdeDF,sourceRows = lahdeRivit)
    write("Saving...")
    moro("Saving...",debug)
    try:
        SOK_Ke.save(
            filename=kohde)
    except PermissionError:
        moro("PermissionError. Kohdetiedosto todennäköisesti auki!")
        write("Lupavirhe! Kohdetiedosto todennäköisesti auki! Tallentaminen epäonnistui.")
    write("Saving process complete.")
    moro("Saving process complete.",debug)



#fuck you! we are in branch firstproto

