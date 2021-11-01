from openpyxl import load_workbook
from openpyxl import utils
import time
from tkinter import *
from tkinter.filedialog import askopenfilename
from Excel.excelExports import ajaArvot, clearWorkbook
from Excel.writers import moro, write
debug = True
# root window/widget
root = Tk()
root.wm_attributes("-transparentcolor", 'grey')
root.title("SOK lomake täyttäjä")
root.geometry("800x520")
sourceExcel = ""
kohdeExcel = ""
lahdeRivit = [1, 1]
kohdeRivit = [28,-1]
pady=2
padx=5
#background photo
bg = PhotoImage(file = "NG58Qz1.png")
label1 = Label( root, image = bg)
label1.place(x = 0, y = 0)
#console

Console = Text(root, height=18)
Console.grid(row=7, column=0,sticky=W, pady=pady, padx=padx, rowspan=8, columnspan=1000)
"""
def write(*message, end = "\n", sep = " "):
    text = ""
    for item in message:
        text += "{}".format(item)
        text += sep
    text += end
    Console.insert(INSERT, text)
"""
# created a label(widget)
labelBg = "#c9c9c9"
myLabel = Label(root, text="Täytä tarvittavat tiedot", bg=labelBg)
lahdeLabel = Label(root, text="Lähde: -", justify=LEFT, bg=labelBg)
kohdeLabel = Label(root, text="Kohde: -", justify=LEFT, bg=labelBg)
# entries
lahdeEntry = Entry(root, width=10, bg="#a6a6a6", borderwidth=3)
kohdeEntry = Entry(root, width=10, bg="#a6a6a6", borderwidth=3)
# putting to screen
myLabel.grid(row=0, column=0,sticky=W, pady=pady, padx=padx)
lahdeLabel.grid(row=2, column=1,sticky=W, pady=pady, padx=padx, columnspan=10)
kohdeLabel.grid(row=3, column=1,sticky=W, pady=pady, padx=padx, columnspan=10)
lahdeEntry.grid(row=4, column=1,sticky=W, pady=pady, padx=padx)
kohdeEntry.grid(row=5, column=1,sticky=W, pady=pady, padx=padx)

# buttons
def lataaLahde():
    global sourceExcel
    filename = askopenfilename()
    sourceExcel = filename
    if(len(filename)>0):
        lahdeLabel.config(text="Lähde: " + sourceExcel)
    moro("source button: "+ sourceExcel,debug)


lahdeButton = Button(root, text="Valitse lähdetiedosto", padx=10, pady=5, command=lataaLahde,width=16)
lahdeButton.grid(row=2, column=0,sticky=W, pady=pady, padx=padx)

def lataaKohde():
    global kohdeExcel
    filename = askopenfilename()
    kohdeExcel = filename
    if (len(filename) > 0):
        kohdeLabel.config(text="Kohde: " + kohdeExcel)
    moro("target button: "+ kohdeExcel,debug)


kohdeButton = Button(root, text="Valitse kohdetiedosto", padx=10, pady=5, command=lataaKohde,width=16)
kohdeButton.grid(row=3, column=0,sticky=W, pady=pady, padx=padx)

def valitseLahdeRivi():
    global lahdeRivit
    row = lahdeEntry.get()
    moro(row,debug)
    # we check entry is a number
    numbers = []
    for word in row.split():
        if word.isdigit():
            numbers.append(int(word))
    if (len(numbers) == 2):
        if(numbers[0]<numbers[1]):
            lahdeRivit = numbers
            lahdeRiviButton.configure(text="Lähderivit: {} - {}".format(lahdeRivit[0], lahdeRivit[1]))


lahdeRiviButton = Button(root, text="Lähderivit", padx=10, pady=5, command=valitseLahdeRivi,width=16)
lahdeRiviButton.grid(row=4, column=0,sticky=W, pady=pady, padx=padx)

def valitseKohdeRivi():
    global kohdeRivit
    row = kohdeEntry.get()
    moro(row,debug)
    # we check entry is a number
    numbers = []
    for word in row.split():
        if word.isdigit():
            numbers.append(int(word))
    if len(numbers) == 1:
        kohdeRivit[0] = numbers[0]
        kohdeRivit[1]=-1
        kohdeRiviButton.configure(text="Kohderivi: {}".format(kohdeRivit[0]))
    elif len(numbers)==2:
        kohdeRivit[0] = numbers[0]
        kohdeRivit[1] = numbers[1]
        kohdeRiviButton.configure(text="Kohderivi: {} - {}".format(kohdeRivit[0], kohdeRivit[1]))
    else:
        kohdeRivit[0] = 32
        kohdeRivit[1] = -1
        kohdeRiviButton.configure(text="Kohderivi")


kohdeRiviButton = Button(root, text="Kohderivi", padx=10, pady=5, command=valitseKohdeRivi,width=16)
kohdeRiviButton.grid(row=5, column=0,sticky=W, pady=pady, padx=padx)

def run():
    global runButton
    global Console
    if (len(sourceExcel)>0)&(len(kohdeExcel)>0)&(sourceExcel!=kohdeExcel)&(kohdeRivit[0]>1)&(lahdeRivit[0]>1):
        try:
            moro("Run button clicked: Writing target workbook",debug)
            ajaArvot(Console,lahdeRivit,kohdeRivit,sourceExcel,kohdeExcel,debug)
        except utils.exceptions.InvalidFileException:
            write("kohdetiedostossa oli jotain vikaa", console=Console)
            write(
            "openpyxl does not support .md file format, please check you can open it with Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm",
            console=Console)
        except ValueError:
            write("Lähdetiedosto ei ole Excel tiedosto",console=Console)
    else: #this is for testing only
        ajaArvot(Console,lahde=sourceExcel,kohde=kohdeExcel, debug=debug)





runButton = Button(root, text="Kirjoita", padx=10, pady=5, command=run,width=8)
runButton.grid(row=6, column=0,sticky=W, pady=pady, padx=padx)


def clear():
    global runButton
    global Console
    if ( (len(kohdeExcel) > 0) & (kohdeRivit[0] < kohdeRivit[1])):
        moro(2, debug)
        write("Tyhjennetään SOK lomaketta", console=Console)
        clearWorkbook(Console, kohdeRivit=kohdeRivit, kohde=kohdeExcel, debug=debug)
    elif ( (len(kohdeExcel) > 0) & (kohdeRivit[0] > kohdeRivit[1])):
        try:
            moro(2, debug)
            write("Tyhjennetään SOK lomaketta", console=Console)
            clearWorkbook(Console, kohdeRivit=[28,1000], kohde=kohdeExcel, debug=debug)
        except utils.exceptions.InvalidFileException:
            write("kohdetiedostossa oli jotain vikaa",console=Console)
            write("openpyxl does not support .md file format, please check you can open it with Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm", console=Console)
    else:  # this is for testing only
        moro("tiedoissa vikaa", debug)
        print(kohdeExcel)
        print(kohdeRivit)
        write("Lähtö tiedoissa on jotain vikaa", console=Console)
        write(str(kohdeRivit[0])+ " "+str(kohdeRivit[1]),console=Console)



clearButton = Button(root, text="Tyhjennä", padx=10, pady=5, command=clear,width=8)
clearButton.grid(row=6, column=1,sticky=W, pady=pady, padx=padx)

root.wm_attributes("-transparentcolor", 'grey')
# main loop
root.mainloop()
