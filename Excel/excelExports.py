from openpyxl import load_workbook
from Excel.writers import printer, clearColumn, moro, write
from Excel.sheeters import writeUutuudet, writeVanhatTuotteet, writeNimet, writeToimitusyks
from tkinter import *
import pandas as pd
#import numpy as np
#from openpyxl import Workbook



# A master function that writes all the values to the target workbook

def ajaArvot(console,lahdeRivit=-1, kohdeRivit=[28,-1], lahde="export_SOK_taulukkovienti_2021-10-27_07-45-20.xlsx", kohde="SOK Käyttötavaroiden erätuotelomake (muut käyttötavarat) v2 33 (version 1) (version 1)_ROSTERi.xlsx",debug=True):
    print(kohdeRivit[0])
    print(type(kohdeRivit[0]))



    moro("Loading source Workbook...", debug)
    lahdeDF = pd.read_excel(lahde)
    moro(lahdeDF,debug=False)
    moro("Loading target Workbook...", debug)
    SOK_Ke = load_workbook(filename=kohde)
    moro(SOK_Ke.sheetnames,debug=False)


    moro("Writing sheets...", debug)
    if lahdeRivit==-1:
        shape = lahdeDF.shape
        lahdeRivit=[2,shape[0]+1]
    #writeVanhatTuotteet(SOK_Ke, targetRow=kohdeRivi, source=lahdeDF,sourceRows = lahdeRivit)
    writeUutuudet(SOK_Ke, targetRows=kohdeRivit, source=lahdeDF,sourceRows = lahdeRivit, console=console)
    #writeToimitusyks(SOK_Ke, targetRows=kohdeRivit, source=lahdeDF,sourceRows = lahdeRivit, console=console)
    #writeNimet(SOK_Ke, targetRows=kohdeRivit, source=lahdeDF,sourceRows = lahdeRivit, console=console)
    moro("Saving...",debug)
    try:
        SOK_Ke.save(
            filename=kohde)
        write("SOK lomake tallennettu", console=console)
        moro("Saving process complete.", debug)
    except PermissionError:
        moro("PermissionError. Kohdetiedosto todennäköisesti auki!")
        write("Lupavirhe! Kohdetiedosto todennäköisesti auki! Tallentaminen epäonnistui.",console=console)

def clearWorkbook(console, kohdeRivit=[28,500], kohde="SOK Käyttötavaroiden erätuotelomake (muut käyttötavarat) v2 33 (version 1) (version 1)_ROSTERi.xlsx",debug=True):
    columns = {'2. Uutuudet - New articles':["P","R","T","V","X","AB","AJ","AH","AM","AO","AQ","AS","BR","BS","BT","BU","BW","CQ","CS","CU","CV","CX","DQ","DS","DW","DY","EA","EE","EG","EI","EK","EM","EO","EQ","ES","EV","EX","EZ","FB","FC"],
               # '2. Toimitusyks. -  Deliv. units':["M","BV","BX","BT","BR","K","AR","AP","AN","AL","I","O",],
               #'3. Nimet - Names':["U","C","M","O"],
               #'5. P.materiaalit - P.materials':["N","R","AO","AQ"],
               #'7. Valmistaja - Manufacturer':["M","AM","AQ",]
               }
    SOK_Ke = load_workbook(filename=kohde)

    for sheetname in columns:
        sheet = SOK_Ke.get_sheet_by_name(sheetname)
        for column in columns[sheetname]:

            clearColumn(sheet,kohdeRivit,column)

    moro("Saving...", debug)
    try:
        SOK_Ke.save(
            filename=kohde)
        write("SOK lomake tallennettu", console=console)
        moro("Saving process complete.", debug)
    except PermissionError:
        moro("PermissionError. Kohdetiedosto todennäköisesti auki!")
        write("Lupavirhe! Kohdetiedosto todennäköisesti auki! Tallentaminen epäonnistui.", console=console)






